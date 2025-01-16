import tkinter as tk
from tkinter import filedialog, messagebox
import speech_recognition as sr
import openpyxl
from tkinter import ttk
from openpyxl.utils import get_column_letter

# Biến toàn cục
recognizer = sr.Recognizer()
audio_data = None

# Hàm để chọn file Excel và hiển thị sheet
def select_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, filepath)
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet_names = wb.sheetnames
            sheet_combo['values'] = sheet_names
            if sheet_names:
                sheet_combo.current(0)  # Chọn sheet đầu tiên
                display_sheet(sheet_names[0], wb)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook: {e}")

# Hiển thị nội dung sheet lên giao diện
def display_sheet(sheet_name, workbook):
    try:
        sheet = workbook[sheet_name]
        for widget in sheet_canvas.winfo_children():
            widget.destroy()

        # Hiển thị tiêu đề cột
        max_col = sheet.max_column
        for j in range(1, max_col + 1):
            column_letter = get_column_letter(j)
            header = tk.Label(sheet_canvas, text=column_letter, bg="#d3d3d3", font=("Arial", 10, "bold"))
            header.grid(row=0, column=j, padx=2, pady=2)

        # Hiển thị dữ liệu
        max_row = sheet.max_row
        for i in range(1, max_row + 1):
            # Hiển thị số hàng
            row_header = tk.Label(sheet_canvas, text=str(i), bg="#d3d3d3", font=("Arial", 10, "bold"))
            row_header.grid(row=i, column=0, padx=2, pady=2)
            for j in range(1, max_col + 1):
                value = sheet.cell(row=i, column=j).value
                cell = tk.Entry(sheet_canvas, width=15, font=("Arial", 10))
                cell.grid(row=i, column=j, padx=2, pady=2)
                cell.insert(0, value if value is not None else "")
                cell.config(state="readonly")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to display sheet: {e}")

# Khi chọn sheet từ combo box
def on_sheet_select(event):
    filepath = file_path_entry.get()
    if not filepath:
        return
    try:
        wb = openpyxl.load_workbook(filepath)
        selected_sheet = sheet_combo.get()
        display_sheet(selected_sheet, wb)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load sheet: {e}")

# Bắt đầu ghi âm
def start_recording():
    global audio_data
    try:
        with sr.Microphone() as source:
            status_label.config(text="Đang lắng nghe...", fg="blue")
            audio_data = recognizer.listen(source, timeout=None)
            status_label.config(text="Đang ghi... Nhấn 'End' để dừng", fg="orange")
    except Exception as e:
        status_label.config(text=f"Error: {e}", fg="red")

# Kết thúc ghi âm và xử lý
def end_recording():
    global audio_data
    try:
        if audio_data is None:
            messagebox.showerror("Error", "No recording started.")
            return
        status_label.config(text="Processing...", fg="orange")
        text = recognizer.recognize_google(audio_data, language="vi-VN")

        # Xử lý thay thế "phẩy" bằng dấu phẩy thực tế
        text = text.replace("phẩy", ",")

        text_entry.delete(0, tk.END)
        text_entry.insert(0, text)
        status_label.config(text="Recognition complete.", fg="green")
    except sr.UnknownValueError:
        status_label.config(text="Could not understand the audio.", fg="red")
    except sr.RequestError as e:
        status_label.config(text=f"Error with the service: {e}", fg="red")

# Hàm để ghi văn bản vào ô Excel
def write_to_excel():
    filepath = file_path_entry.get()
    if not filepath:
        messagebox.showerror("Error", "Please select an Excel file.")
        return

    text = text_entry.get()
    if not text:
        messagebox.showerror("Error", "No text to write.")
        return

    sheet_name = sheet_combo.get()
    cell = cell_entry.get()

    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{sheet_name}' not found in the workbook.")
            return

        sheet = wb[sheet_name]
        sheet[cell] = text
        wb.save(filepath)
        messagebox.showinfo("Success", f"Text written to {cell} in sheet '{sheet_name}'.")

        # Load lại nội dung sheet sau khi ghi
        display_sheet(sheet_name, wb)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to write to Excel: {e}")

# Tạo giao diện chính
root = tk.Tk()
root.title("Phần mềm chuyển đổi giọng nói")
root.geometry("1000x800")
root.resizable(True, True)
root.configure(bg="#f4f4f4")

# Header
header_label = tk.Label(root, text="Phần mềm chuyển đổi giọng nói", font=("Arial", 16, "bold"), bg="#f4f4f4", fg="#333")
design_by = tk.Label(root, text="Design by: Thịnh Nguyễn (GĐYK)", font=("Arial", 10, "bold"), bg="#f4f4f4", fg="#333")
header_label.pack(pady=(10, 0))
design_by.pack(pady=(0, 20))

# Frame chọn file
file_frame = tk.Frame(root, bg="#f4f4f4")
file_frame.pack(pady=5)
tk.Label(file_frame, text="Chọn File Excel:", font=("Arial", 10), bg="#f4f4f4").grid(row=0, column=0, sticky="w", padx=5)
file_path_entry = tk.Entry(file_frame, width=30, font=("Arial", 10))
file_path_entry.grid(row=0, column=1, padx=5)
file_button = tk.Button(file_frame, text="Browse", command=select_file, bg="#0078D7", fg="white", font=("Arial", 10))
file_button.grid(row=0, column=2, padx=5)

# Combo box chọn sheet
tk.Label(file_frame, text="Chọn Sheet:", font=("Arial", 10), bg="#f4f4f4").grid(row=1, column=0, sticky="w", padx=5)
sheet_combo = ttk.Combobox(file_frame, font=("Arial", 10))
sheet_combo.grid(row=1, column=1, padx=5)
sheet_combo.bind("<<ComboboxSelected>>", on_sheet_select)

# Frame cuộn cho sheet
sheet_frame = tk.Frame(root, bg="#f4f4f4")
sheet_frame.pack(pady=10, fill=tk.BOTH, expand=True)
sheet_canvas = tk.Canvas(sheet_frame, bg="#f4f4f4")
sheet_scroll = tk.Scrollbar(sheet_frame, orient=tk.VERTICAL, command=sheet_canvas.yview)
sheet_scroll.pack(side=tk.RIGHT, fill=tk.Y)
sheet_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
sheet_canvas.configure(yscrollcommand=sheet_scroll.set)

# Frame hiển thị sheet trong canvas
sheet_inner_frame = tk.Frame(sheet_canvas, bg="#f4f4f4")
sheet_canvas.create_window((0, 0), window=sheet_inner_frame, anchor="nw")
sheet_canvas.bind("<Configure>", lambda e: sheet_canvas.configure(scrollregion=sheet_canvas.bbox("all")))

# Frame nhập ô và ghi âm
input_frame = tk.Frame(root, bg="#f4f4f4")
input_frame.pack(pady=5)
tk.Label(input_frame, text="Cell (e.g., A1):", font=("Arial", 10), bg="#f4f4f4").grid(row=0, column=0, sticky="w", padx=5)
cell_entry = tk.Entry(input_frame, font=("Arial", 10))
cell_entry.grid(row=0, column=1, padx=5)

# Frame chức năng ghi âm
record_frame = tk.Frame(root, bg="#f4f4f4")
record_frame.pack(pady=10)
record_button = tk.Button(record_frame, text="Bắt đầu ghi", command=start_recording, bg="#28A745", fg="white", font=("Arial", 10))
record_button.grid(row=0, column=0, padx=10)
end_button = tk.Button(record_frame, text="Kết thúc", command=end_recording, bg="#DC3545", fg="white", font=("Arial", 10))
end_button.grid(row=0, column=1, padx=10)

# Frame hiển thị văn bản
text_frame = tk.Frame(root, bg="#f4f4f4")
text_frame.pack(pady=5)
tk.Label(text_frame, text="Văn bản:", font=("Arial", 10), bg="#f4f4f4").grid(row=0, column=0, sticky="w", padx=5)
text_entry = tk.Entry(text_frame, width=40, font=("Arial", 10))
text_entry.grid(row=0, column=1, padx=5)

# Nút ghi vào Excel
write_button = tk.Button(root, text="Lưu Excel", command=write_to_excel, bg="#0078D7", fg="white", font=("Arial", 10))
write_button.pack(pady=10)

# Status label
status_label = tk.Label(root, text="", font=("Arial", 10), bg="#f4f4f4", fg="green")
status_label.pack(pady=5)

# Chạy giao diện
root.mainloop()
